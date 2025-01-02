import os # Modul untuk berinteraksi dengan sistem file
import pandas as pd  # Pustaka untuk analisis data dan manipulasi tabel
from tabulate import tabulate # Untuk membuat tabel dalam format teks
from openpyxl import Workbook, load_workbook  # Untuk membuat dan memodifikasi file Excel
from openpyxl.utils import get_column_letter  # Untuk mengonversi nomor kolom menjadi huruf
from openpyxl.styles import Alignment, Font  # Untuk memformat teks dalam file Excel


# Direktori penyimpanan
ATTENDANCE_FILE = "./data/attendance/attendance.csv"  # File CSV untuk mencatat kehadiran
ATTENDANCE_FILE_XLSX = "./data/attendance_xlsx/attendance.xlsx"  # File Excel untuk mencatat kehadiran

# Fungsi untuk menampilkan attendance.csv dalam bentuk tabel
def display_attendance():
    # Mengecek apakah file ATTENDANCE_FILE (CSV) ada
    if not os.path.exists(ATTENDANCE_FILE):
        print("Attendance file does not exist.")  # Menampilkan pesan jika file tidak ditemukan
        return  # Keluar dari fungsi

    # Membaca file CSV ke dalam DataFrame menggunakan pandas
    df = pd.read_csv(ATTENDANCE_FILE)

    # Menampilkan data dalam bentuk tabel di terminal
    print("\n--- Attendance Records ---")
    # Menggunakan tabulate untuk membuat tampilan tabel yang rapi
    print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))

# Fungsi untuk membuka file attendance.csv dalam format Excel
def open_attendance_with_excel():
    try:
        # Mendefinisikan file CSV yang akan dibuka
        csv_file = ATTENDANCE_FILE

        # Mengecek apakah file CSV ada
        if not os.path.exists(csv_file):
            print(f"File '{csv_file}' tidak ditemukan.")  # Menampilkan pesan jika file tidak ditemukan
            return  # Keluar dari fungsi

        # Membaca file CSV ke dalam DataFrame menggunakan pandas
        df = pd.read_csv(csv_file)

        # Mendefinisikan file Excel output
        excel_file = ATTENDANCE_FILE_XLSX

        # Menyimpan DataFrame ke dalam file Excel menggunakan openpyxl
        df.to_excel(excel_file, index=False, engine='openpyxl')

        # Membuka file Excel untuk modifikasi (auto-width kolom)
        wb = load_workbook(excel_file)  # Membuka workbook Excel
        ws = wb.active  # Mengakses sheet aktif

        # Terapkan auto-width untuk semua kolom
        for col in ws.columns:  # Iterasi untuk setiap kolom
            max_length = 0  # Menyimpan panjang maksimum dalam kolom
            column = col[0].column  # Nomor kolom (1-based index)
            for cell in col:  # Iterasi untuk setiap sel dalam kolom
                try:
                    # Mengukur panjang isi sel jika ada
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Mengatur lebar kolom dengan padding tambahan
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Format header dengan teks tebal dan rata tengah
        for cell in ws[1]:  # Baris pertama dianggap header
            cell.font = Font(bold=True)  # Membuat teks tebal
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Rata tengah

        # Menyimpan perubahan ke file Excel
        wb.save(excel_file)

        print(f"File '{excel_file}' berhasil dibuat dengan kolom rapi.")  # Pesan keberhasilan

        # Membuka file Excel menggunakan perintah sistem (hanya Windows)
        os.system(f'start EXCEL.EXE "{excel_file}"')

    except Exception as e:
        # Menangkap dan menampilkan kesalahan jika terjadi
        print(f"Terjadi kesalahan: {e}")