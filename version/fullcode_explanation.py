# Mengimpor pustaka-pustaka yang diperlukan

import cv2  # OpenCV, pustaka untuk pemrosesan gambar dan video
import os  # Modul untuk berinteraksi dengan sistem file
import csv  # Modul untuk membaca dan menulis file CSV
import numpy as np  # Pustaka untuk operasi numerik dan array
import pandas as pd  # Pustaka untuk analisis data dan manipulasi tabel
import sys  # Modul untuk fungsi dan objek sistem
import pyttsx3  # Pustaka untuk teks ke suara (text-to-speech)
import speech_recognition as sr  # Pustaka untuk pengenalan suara
import skfuzzy as fuzz  # Pustaka untuk logika fuzzy
import skfuzzy.control as ctrl  # Modul kontrol untuk logika fuzzy
import smtplib  # Library untuk mengirim email menggunakan protokol SMTP

from email.mime.multipart import MIMEMultipart  # Untuk membuat email dengan banyak bagian (subject, body, dll.)
from email.mime.text import MIMEText  # Untuk membuat bagian teks email
from tabulate import tabulate  # Untuk membuat tabel dalam format teks
from datetime import datetime  # Untuk bekerja dengan tanggal dan waktu
from openpyxl import Workbook, load_workbook  # Untuk membuat dan memodifikasi file Excel
from openpyxl.utils import get_column_letter  # Untuk mengonversi nomor kolom menjadi huruf
from openpyxl.styles import Alignment, Font  # Untuk memformat teks dalam file Excel

# Direktori penyimpanan
DATASET_DIR = "./data/face_dataset"  # Direktori untuk menyimpan dataset wajah
MODEL_DIR = "./data/trained_model"  # Direktori untuk menyimpan model yang telah dilatih

ATTENDANCE_FILE = "./data/attendance/attendance.csv"  # File CSV untuk mencatat kehadiran
ATTENDANCE_FILE_XLSX = "./data/attendance_xlsx/attendance.xlsx"  # File Excel untuk mencatat kehadiran
USERS_FILE = "./data/users/users.csv"  # File CSV untuk menyimpan data pengguna

# Inisialisasi Haar Cascades dan Recognizer
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")  
# Detektor wajah menggunakan Haar Cascade bawaan OpenCV

eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye.xml")  
# Detektor mata menggunakan Haar Cascade bawaan OpenCV

smile_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_smile.xml")  
# Detektor senyuman menggunakan Haar Cascade bawaan OpenCV

recognizer = cv2.face.LBPHFaceRecognizer_create()  
# Inisialisasi Local Binary Pattern Histogram (LBPH) Recognizer untuk pengenalan wajah

# Variabel untuk deteksi
EYE_CLOSED_FRAMES = 2  # Jumlah frame berturut-turut mata tertutup yang dianggap sebagai satu kedipan
eye_blink_count = 0  # Counter untuk jumlah kedipan mata yang terdeteksi
blink_counter = 0  # Counter untuk menghitung jumlah frame mata tertutup berturut-turut
smile_detected = False  # Flag untuk mendeteksi apakah senyuman ditemukan

# Inisialisasi Text-to-Speech
engine = pyttsx3.init()  # Menginisialisasi engine Text-to-Speech
engine.setProperty('rate', 150)  # Mengatur kecepatan bicara (kata per menit)
engine.setProperty('volume', 0.9)  # Mengatur volume suara (0.0 hingga 1.0)

# Fungsi untuk mengucapkan teks
def speak_text(text):
    engine.say(text)  # Menggunakan engine Text-to-Speech untuk mengucapkan teks
    engine.runAndWait()  # Menjalankan proses Text-to-Speech hingga selesai

# Fungsi untuk mendengarkan dan mengenali ucapan
def listen_for_phrase():
    recognizer = sr.Recognizer()  # Membuat objek recognizer untuk mengenali suara
    with sr.Microphone() as source:  # Menggunakan mikrofon sebagai sumber input suara
        print("Listening for 'I am here'...")  # Memberi tahu pengguna bahwa aplikasi sedang mendengarkan
        try:
            recognizer.adjust_for_ambient_noise(source, duration=1)  
            # Menyesuaikan sensitivitas deteksi suara berdasarkan kebisingan lingkungan selama 1 detik
            audio = recognizer.listen(source, timeout=10)  
            # Mendengarkan input suara dengan batas waktu 10 detik
            recognized_text = recognizer.recognize_google(audio)  
            # Menggunakan API Google untuk mengenali teks dari audio
            print(f"Recognized: {recognized_text}")  
            # Menampilkan teks yang dikenali ke layar
            return recognized_text.lower()  
            # Mengembalikan teks yang dikenali dalam huruf kecil untuk mempermudah perbandingan
        except sr.UnknownValueError:
            # Jika tidak dapat mengenali ucapan
            print("Could not understand the audio.")
        except sr.RequestError as e:
            # Jika ada kesalahan dalam permintaan ke layanan pengenalan ucapan
            print(f"Could not request results; {e}")
        except sr.WaitTimeoutError:
            # Jika waktu mendengarkan habis tanpa menerima suara
            print("Listening timed out.")
    return None  # Mengembalikan None jika tidak ada teks yang dikenali

# Fungsi untuk mendapatkan informasi pengguna berdasarkan ID
def get_user_info_by_id(user_id):
    users_file = USERS_FILE  # Lokasi file pengguna
    if not os.path.exists(users_file):  # Periksa apakah file pengguna ada
        return None  # Jika tidak ada, kembalikan None

    with open(users_file, "r") as f:  # Buka file pengguna dalam mode baca
        reader = csv.DictReader(f)  # Membaca file sebagai dictionary
        for row in reader:  # Iterasi setiap baris dalam file
            if row["ID"] == user_id:  # Jika ID cocok dengan user_id
                return row["Name"], row["Email"]  # Kembalikan nama dan email
    return None  # Jika tidak ditemukan, kembalikan None

# Fuzzy Logic untuk deteksi kedipan mata
def fuzzy_eye_blink(eyes_detected, eye_closed_frames):
    # Variabel fuzzy
    blink_duration = ctrl.Antecedent(np.arange(0, 10, 1), 'blink_duration')  # Durasi kedipan dalam frame
    blink_confidence = ctrl.Consequent(np.arange(0, 1.1, 0.1), 'blink_confidence')  # Kepercayaan kedipan

    # Membership functions
    blink_duration['short'] = fuzz.trapmf(blink_duration.universe, [0, 0, 2, 4])
    blink_duration['medium'] = fuzz.trimf(blink_duration.universe, [2, 4, 6])
    blink_duration['long'] = fuzz.trapmf(blink_duration.universe, [4, 6, 10, 10])

    blink_confidence['low'] = fuzz.trapmf(blink_confidence.universe, [0, 0, 0.3, 0.5])
    blink_confidence['medium'] = fuzz.trimf(blink_confidence.universe, [0.3, 0.5, 0.7])
    blink_confidence['high'] = fuzz.trapmf(blink_confidence.universe, [0.5, 0.7, 1.0, 1.0])

    # Aturan fuzzy
    rules = [
        ctrl.Rule(blink_duration['short'], blink_confidence['low']),
        ctrl.Rule(blink_duration['medium'], blink_confidence['medium']),
        ctrl.Rule(blink_duration['long'], blink_confidence['high']),
    ]

    # Sistem kontrol fuzzy
    system = ctrl.ControlSystem(rules)  # Inisialisasi sistem kontrol fuzzy
    simulation = ctrl.ControlSystemSimulation(system)  # Simulasi sistem kontrol

    # Hitung kepercayaan berdasarkan input
    simulation.input['blink_duration'] = blink_counter  # Masukkan jumlah frame kedipan
    simulation.compute()  # Jalankan simulasi fuzzy
    confidence = simulation.output['blink_confidence']  # Dapatkan output kepercayaan

    return confidence  # Kembalikan kepercayaan

# Panggil fungsi fuzzy dalam pendeteksian kedipan mata
def check_eye_blink(eyes_detected, eye_closed_frames):
    global eye_blink_count, blink_counter
    if eyes_detected == 0:  # Jika mata tertutup
        blink_counter += 1  # Tambahkan counter kedipan
    else:  # Jika mata terbuka
        if blink_counter >= eye_closed_frames:  # Jika jumlah frame mata tertutup melewati ambang batas
            confidence = fuzzy_eye_blink(eyes_detected, eye_closed_frames)  # Hitung kepercayaan fuzzy
            if confidence > 0.5:  # Jika kepercayaan melebihi ambang batas
                eye_blink_count += 1  # Tambahkan hitungan kedipan valid
        blink_counter = 0  # Reset counter setelah mata terbuka

    return eye_blink_count  # Kembalikan jumlah kedipan

# Fungsi fuzzy untuk memeriksa senyuman berdasarkan ukuran senyuman
def fuzzy_smile_check(smiles, face_width):
    if len(smiles) > 0:  # Periksa jika ada senyuman yang terdeteksi
        smile_ratio = smiles[0][2] / face_width  # Rasio lebar senyuman terhadap lebar wajah
    else:
        smile_ratio = 0  # Jika tidak ada senyuman, rasio 0

    # Variabel fuzzy
    smile_size = ctrl.Antecedent(np.arange(0, 1.1, 0.1), 'smile_size')  # Ukuran senyuman relatif
    smile_confidence = ctrl.Consequent(np.arange(0, 1.1, 0.1), 'smile_confidence')  # Kepercayaan senyuman

    # Membership functions
    smile_size['small'] = fuzz.trapmf(smile_size.universe, [0, 0, 0.3, 0.5])
    smile_size['medium'] = fuzz.trimf(smile_size.universe, [0.3, 0.5, 0.7])
    smile_size['large'] = fuzz.trapmf(smile_size.universe, [0.5, 0.7, 1.0, 1.0])

    smile_confidence['low'] = fuzz.trapmf(smile_confidence.universe, [0, 0, 0.3, 0.5])
    smile_confidence['medium'] = fuzz.trimf(smile_confidence.universe, [0.3, 0.5, 0.7])
    smile_confidence['high'] = fuzz.trapmf(smile_confidence.universe, [0.5, 0.7, 1.0, 1.0])

    # Aturan fuzzy
    rules = [
        ctrl.Rule(smile_size['small'], smile_confidence['low']),
        ctrl.Rule(smile_size['medium'], smile_confidence['medium']),
        ctrl.Rule(smile_size['large'], smile_confidence['high']),
    ]

    # Sistem kontrol fuzzy
    system = ctrl.ControlSystem(rules)  # Inisialisasi sistem kontrol fuzzy
    simulation = ctrl.ControlSystemSimulation(system)  # Simulasi sistem kontrol

    # Hitung kepercayaan berdasarkan input
    simulation.input['smile_size'] = smile_ratio  # Masukkan rasio senyuman
    simulation.compute()  # Jalankan simulasi fuzzy
    confidence = simulation.output['smile_confidence']  # Dapatkan output kepercayaan

    return confidence > 0.5  # Kembalikan True jika kepercayaan melebihi ambang batas

# Fungsi utama untuk deteksi wajah, kedipan mata, dan senyuman
def recognize_faces_and_attendance():
    # Fungsi utama untuk deteksi wajah, kedipan mata, dan senyuman, serta mencatat kehadiran.

    global eye_blink_count, smile_detected, blink_counter  # Gunakan variabel global untuk kedipan dan senyuman.

    # Reset variabel untuk sesi baru
    eye_blink_count = 0  # Mengatur ulang jumlah kedipan mata.
    smile_detected = False  # Mengatur ulang status senyuman.
    blink_counter = 0  # Mengatur ulang counter kedipan berturut-turut.

    recognizer.read(os.path.join(MODEL_DIR, "face_recognizer.yml"))  # Memuat model pengenalan wajah dari file.

    # Tampilkan instruksi kepada pengguna untuk memverifikasi kehadiran
    print("please blink 3x then smile for first verification!")
    speak_text("please blink 3x then smile for first verification!")  # Menggunakan text-to-speech untuk instruksi.

    # Load label map untuk menerjemahkan ID ke nama pengguna.
    label_map = {}
    with open(os.path.join(MODEL_DIR, "labels.txt"), "r") as f:
        for line in f:
            label, user_id = line.strip().split(",")  # Membaca label dan ID dari file.
            label_map[int(label)] = user_id  # Menyimpan ke dalam dictionary.

    cap = cv2.VideoCapture(0)  # Mengaktifkan kamera.
    print("Press 'q' to quit.")  # Petunjuk untuk keluar dari aplikasi.

    while cap.isOpened():  # Loop selama kamera terbuka.
        ret, frame = cap.read()  # Membaca frame dari kamera.
        if not ret:  # Jika gagal membaca frame, keluar dari loop.
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)  # Mengonversi frame ke grayscale.
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)  # Mendeteksi wajah dalam frame.

        for (x, y, w, h) in faces:  # Iterasi melalui setiap wajah yang terdeteksi.
            face_resized = cv2.resize(gray[y:y + h, x:x + w], (200, 200))  # Memperbesar ukuran wajah untuk prediksi.
            try:
                # Memprediksi wajah menggunakan recognizer.
                label, confidence = recognizer.predict(face_resized)
                if confidence > 50.0:  # Jika confidence rendah, wajah dianggap "Unknown".
                    detected_id = "Unknown"
                else:
                    detected_id = label_map.get(label, "Unknown")  # Mencocokkan label dengan ID pengguna.
            except Exception as e:
                detected_id = "Unknown"  # Jika terjadi error, wajah dianggap "Unknown".

            # Ambil data pengguna berdasarkan ID.
            user_info = get_user_info_by_id(detected_id)
            if user_info:  # Jika data pengguna ditemukan.
                user_name, email = user_info  # Mengambil nama dan email.
                display_text = f"{detected_id} | {user_name}"  # Format teks untuk ditampilkan.
            else:
                display_text = f"Unknown"  # Tampilkan sebagai "Unknown" jika data tidak ditemukan.

            # Tampilkan persegi wajah dan informasi pengguna di layar.
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 255), 2)  # Gambar kotak di sekitar wajah.
            cv2.putText(frame, display_text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, 
                        (0, 255, 0) if detected_id != "Unknown" else (0, 0, 255), 2)
            if detected_id != "Unknown":
                cv2.putText(frame, f"Confidence: {confidence:.2f}%", (x, y + h + 20), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            # Pemeriksaan tambahan untuk kedipan mata dan senyuman jika wajah dikenali.
            if detected_id != "Unknown":
                roi_gray = gray[y:y + h, x:x + w]  # Region of Interest (ROI) untuk wajah.

                # Deteksi mata.
                eyes = eye_cascade.detectMultiScale(roi_gray, 1.1, 10, minSize=(30, 30))
                current_blink_count = check_eye_blink(len(eyes), EYE_CLOSED_FRAMES)  # Periksa kedipan mata.
                cv2.putText(frame, f"Eye Blink Count: {current_blink_count}", (x, y - 40), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

                # Deteksi senyuman.
                smiles = smile_cascade.detectMultiScale(roi_gray, 1.8, 20, minSize=(30, 30))
                smile_detected = fuzzy_smile_check(smiles, w)  # Gunakan fuzzy logic untuk memeriksa senyuman.
                if smile_detected:  # Jika senyuman terdeteksi.
                    cv2.putText(frame, "Toothy Smile: Detected", (x, y - 70), 
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

                # Logika verifikasi kehadiran.
                if current_blink_count >= 3 and smile_detected:  # Jika kedipan dan senyuman terdeteksi.
                    print("Please say 'I am here' to confirm your attendance.")
                    speak_text("Please say 'I am here' to confirm your attendance.")  # Instruksi konfirmasi.
                    response = listen_for_phrase()  # Dengarkan respons pengguna.
                    if response and "i am here" in response:  # Jika respons valid.
                        cv2.putText(frame, "Verified", (x + w - 30, y + 30), 
                                    cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 3)

                        # Catat kehadiran ke file.
                        with open(ATTENDANCE_FILE, "a", newline="") as file:
                            writer = csv.writer(file)
                            writer.writerow([detected_id, user_name, "Hadir", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                        print(f"{user_name} (ID: {detected_id}) marked as present.")

                        # Kirim email notifikasi.
                        email_subject = "Kehadiran Anda Tercatat"
                        email_body = (
                            f"Halo {user_name},\n\n"
                            f"Kehadiran Anda telah berhasil dicatat pada {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n\n"
                            f"ID: {detected_id}\n\n"
                            f"Terima kasih."
                        )
                        send_email_notification(email, email_subject, email_body)  # Kirim email.

                        print("Your attendance has been recorded, please check your email to see it.")
                        speak_text("Your attendance has been recorded, please check your email to see it.")

                        cap.release()  # Lepaskan kamera.
                        cv2.destroyAllWindows()  # Tutup semua jendela OpenCV.
                        return  # Keluar setelah kehadiran tercatat.
                    else:
                        print("Kehadiran Anda gagal dicatat. Silakan coba lagi.")
                        cap.release()
                        cv2.destroyAllWindows()
                        return  # Keluar jika gagal mencatat kehadiran.

        # Tampilkan frame di jendela OpenCV.
        cv2.imshow("Smart Attendance System", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):  # Tekan 'q' untuk keluar.
            break

    cap.release()  # Lepaskan kamera.
    cv2.destroyAllWindows()  # Tutup semua jendela OpenCV.

def collect_images():
    # Fungsi untuk mengumpulkan gambar wajah dan menyimpannya di dataset lokal.

    # Input nama, ID, dan email di dalam fungsi.
    person_id = input("Enter the person's ID (NIP/NIM): ").strip()  # Meminta input ID pengguna.
    person_name = input("Enter the person's name: ").strip()  # Meminta input nama pengguna.
    email = input("Enter the person's Email: ").strip()  # Meminta input email pengguna.

    # Buat direktori berdasarkan ID pengguna.
    cap = cv2.VideoCapture(0)  # Mengaktifkan kamera untuk pengumpulan gambar.
    count = 0  # Counter untuk menghitung jumlah gambar yang berhasil diambil.
    person_dir = os.path.join(DATASET_DIR, person_id)  # Direktori penyimpanan gambar pengguna berdasarkan ID.
    os.makedirs(person_dir, exist_ok=True)  # Membuat direktori jika belum ada.

    # Simpan data pengguna ke dalam `users.csv`.
    users_file = USERS_FILE  # Lokasi file penyimpanan data pengguna.
    if not os.path.exists(users_file):  # Jika file belum ada, buat file baru dengan header.
        with open(users_file, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Email"])  # Menulis header ke file CSV.

    # Tambahkan data pengguna ke file.
    with open(users_file, "a", newline="") as f:  # Membuka file CSV dalam mode append.
        writer = csv.writer(f)
        writer.writerow([person_id, person_name, email])  # Menambahkan data pengguna ke file.

    # Mulai pengumpulan gambar.
    print(f"Collecting images for '{person_name}' (ID: {person_id}). Press 'q' to quit early.")  
    # Memberikan informasi kepada pengguna tentang proses pengumpulan gambar.
    
    while cap.isOpened() and count < 100:  # Mengambil hingga 100 gambar atau sampai 'q' ditekan.
        ret, frame = cap.read()  # Membaca frame dari kamera.
        if not ret:  # Jika gagal membaca frame, keluar dari loop.
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)  # Mengonversi frame ke grayscale.
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)  # Mendeteksi wajah pada frame.

        for (x, y, w, h) in faces:  # Iterasi melalui setiap wajah yang terdeteksi.
            face = gray[y:y + h, x:x + w]  # Memotong ROI (Region of Interest) wajah.
            face_resized = cv2.resize(face, (200, 200))  # Mengubah ukuran wajah ke 200x200 piksel.
            file_path = os.path.join(person_dir, f"{count}.jpg")  # Path penyimpanan file gambar.
            cv2.imwrite(file_path, face_resized)  # Menyimpan gambar wajah ke file.
            count += 1  # Increment counter setelah berhasil menyimpan gambar.

            # Menampilkan progress pengumpulan gambar di layar.
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)  # Gambar kotak di sekitar wajah.
            cv2.putText(frame, f"Image {count}/100", (10, 30), 
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)  # Tampilkan jumlah gambar yang terkumpul.

        cv2.imshow("Collecting Faces", frame)  # Menampilkan frame di jendela OpenCV.
        if cv2.waitKey(1) & 0xFF == ord('q'):  # Keluar jika tombol 'q' ditekan.
            break

    cap.release()  # Melepaskan kamera setelah selesai.
    cv2.destroyAllWindows()  # Menutup semua jendela OpenCV.
    print(f"Collected {count} images for '{person_name}' (ID: {person_id}').")  
    # Memberikan laporan akhir tentang jumlah gambar yang berhasil dikumpulkan.

def train_recognizer():
    # Fungsi untuk melatih model pengenalan wajah menggunakan data gambar yang telah dikumpulkan.

    faces, labels = [], []  # List untuk menyimpan gambar wajah dan label masing-masing.
    label_map = {}  # Peta label untuk menghubungkan label numerik dengan ID pengguna.

    print("Training the recognizer...")  # Memberikan informasi bahwa proses pelatihan sedang dimulai.

    # Iterasi melalui setiap direktori pengguna di folder dataset.
    for label, user_id in enumerate(os.listdir(DATASET_DIR)):  
        # Enumerasi memberikan label numerik untuk setiap ID pengguna.
        label_map[label] = user_id  # Menyimpan hubungan label numerik dengan ID pengguna.
        user_dir = os.path.join(DATASET_DIR, user_id)  # Mendapatkan path direktori pengguna.

        # Iterasi melalui setiap file gambar di direktori pengguna.
        for file_name in os.listdir(user_dir):
            img_path = os.path.join(user_dir, file_name)  # Path gambar wajah.
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)  # Membaca gambar dalam mode grayscale.
            faces.append(img)  # Menambahkan gambar ke dalam daftar `faces`.
            labels.append(label)  # Menambahkan label numerik ke daftar `labels`.

    # Melatih model pengenalan wajah menggunakan gambar dan label yang dikumpulkan.
    recognizer.train(faces, np.array(labels))  
    # Menyimpan model yang telah dilatih ke file `face_recognizer.yml`.
    recognizer.write(os.path.join(MODEL_DIR, "face_recognizer.yml"))

    # Menyimpan peta label ke file `labels.txt`.
    with open(os.path.join(MODEL_DIR, "labels.txt"), "w") as f:
        for label, user_id in label_map.items():  
            # Menyimpan setiap pasangan label dan ID pengguna ke dalam file.
            f.write(f"{label},{user_id}\n")  

    print("Model trained and saved successfully.")  # Memberikan informasi bahwa pelatihan selesai dan model telah disimpan.
    
# Fungsi untuk menampilkan attendance.csv dalam bentuk tabel
def display_attendance():
    # Mengecek apakah file ATTENDANCE_FILE (CSV) ada
    if not os.path.exists(ATTENDANCE_FILE):
        print("Attendance file does not exist.")  # Menampilkan pesan jika file tidak ditemukan
        return  # Keluar dari fungsi

    # Membaca file CSV ke dalam DataFrame menggunakan pandas
    df = pd.read_csv(ATTENDANCE_FILE)

    # Menampilkan data dalam bentuk tabel di terminal
    print("\n--- Attendance Records ---")
    # Menggunakan tabulate untuk membuat tampilan tabel yang rapi
    print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))

# Fungsi untuk membuka file attendance.csv dalam format Excel
def open_attendance_with_excel():
    try:
        # Mendefinisikan file CSV yang akan dibuka
        csv_file = ATTENDANCE_FILE

        # Mengecek apakah file CSV ada
        if not os.path.exists(csv_file):
            print(f"File '{csv_file}' tidak ditemukan.")  # Menampilkan pesan jika file tidak ditemukan
            return  # Keluar dari fungsi

        # Membaca file CSV ke dalam DataFrame menggunakan pandas
        df = pd.read_csv(csv_file)

        # Mendefinisikan file Excel output
        excel_file = ATTENDANCE_FILE_XLSX

        # Menyimpan DataFrame ke dalam file Excel menggunakan openpyxl
        df.to_excel(excel_file, index=False, engine='openpyxl')

        # Membuka file Excel untuk modifikasi (auto-width kolom)
        wb = load_workbook(excel_file)  # Membuka workbook Excel
        ws = wb.active  # Mengakses sheet aktif

        # Terapkan auto-width untuk semua kolom
        for col in ws.columns:  # Iterasi untuk setiap kolom
            max_length = 0  # Menyimpan panjang maksimum dalam kolom
            column = col[0].column  # Nomor kolom (1-based index)
            for cell in col:  # Iterasi untuk setiap sel dalam kolom
                try:
                    # Mengukur panjang isi sel jika ada
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Mengatur lebar kolom dengan padding tambahan
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Format header dengan teks tebal dan rata tengah
        for cell in ws[1]:  # Baris pertama dianggap header
            cell.font = Font(bold=True)  # Membuat teks tebal
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Rata tengah

        # Menyimpan perubahan ke file Excel
        wb.save(excel_file)

        print(f"File '{excel_file}' berhasil dibuat dengan kolom rapi.")  # Pesan keberhasilan

        # Membuka file Excel menggunakan perintah sistem (hanya Windows)
        os.system(f'start EXCEL.EXE "{excel_file}"')

    except Exception as e:
        # Menangkap dan menampilkan kesalahan jika terjadi
        print(f"Terjadi kesalahan: {e}")


# Fungsi untuk mengirim notifikasi email
def send_email_notification(to_email, subject, body):
    # Konfigurasi email
    from_email = "abdulrohmanm1304@gmail.com"  # Alamat email pengirim
    password = "wzqz ijmp mbyw hnmb"  # Password aplikasi email pengirim

    try:
        # Setup email server
        server = smtplib.SMTP('smtp.gmail.com', 587)  # Menggunakan server SMTP Gmail pada port 587
        server.starttls()  # Memulai koneksi TLS (untuk keamanan)
        server.login(from_email, password)  # Login ke akun email menggunakan kredensial

        # Buat pesan email
        msg = MIMEMultipart()  # Membuat objek email multipart
        msg['From'] = from_email  # Menentukan alamat pengirim
        msg['To'] = to_email  # Menentukan alamat penerima
        msg['Subject'] = subject  # Menentukan subjek email
        msg.attach(MIMEText(body, 'plain'))  # Menambahkan isi email dengan format teks biasa

        # Kirim email
        server.send_message(msg)  # Mengirim email
        print(f"Email berhasil dikirim ke {to_email}")  # Konfirmasi bahwa email berhasil dikirim
        server.quit()  # Menutup koneksi ke server email
    except Exception as e:
        # Jika ada kesalahan, tangkap dan tampilkan pesan kesalahan
        print(f"Gagal mengirim email: {e}")
        
# Menu utama program
if __name__ == "__main__":  # Mengecek apakah file ini dijalankan langsung (bukan diimpor sebagai modul)
    if len(sys.argv) > 1:  # Mengecek apakah ada argumen tambahan yang diberikan saat menjalankan skrip
        action = sys.argv[1]  # Mengambil argumen pertama setelah nama file
        # Mengecek nilai argumen untuk menentukan tindakan
        if action == "collect_images":
            collect_images()  # Memanggil fungsi untuk mengumpulkan data wajah
        elif action == "train_recognizer":
            train_recognizer()  # Memanggil fungsi untuk melatih model pengenal wajah
        elif action == "run_attendance":
            recognize_faces_and_attendance()  # Memanggil fungsi untuk menjalankan sistem absensi
        elif action == "view_attendance":
            display_attendance()  # Memanggil fungsi untuk menampilkan data absensi di terminal
        elif action == "open_excel":
            open_attendance_with_excel()  # Memanggil fungsi untuk membuka data absensi menggunakan Excel
    else:
        # Jika tidak ada argumen tambahan, gunakan pendekatan berbasis menu interaktif
        while True:  # Loop tak terbatas sampai pengguna memilih keluar
            print("\n--- Smart Attendance System ---")  # Menampilkan judul menu
            print("1: Collect Face Data")  # Opsi untuk mengumpulkan data wajah
            print("2: Train Recognizer")  # Opsi untuk melatih model pengenal wajah
            print("3: Run Attendance System")  # Opsi untuk menjalankan sistem absensi
            print("4: View Attendance Records in Terminal")  # Opsi untuk melihat data absensi di terminal
            print("5: Open Attendance Records in Excel")  # Opsi untuk membuka data absensi menggunakan Excel
            print("6: Exit")  # Opsi untuk keluar dari program
            choice = input("Enter your choice: ")  # Meminta input pengguna untuk memilih opsi menu

            # Mengecek input pengguna dan menjalankan fungsi yang sesuai
            if choice == "1":
                collect_images()  # Memanggil fungsi untuk mengumpulkan data wajah
            elif choice == "2":
                train_recognizer()  # Memanggil fungsi untuk melatih model pengenal wajah
            elif choice == "3":
                recognize_faces_and_attendance()  # Memanggil fungsi untuk menjalankan sistem absensi
            elif choice == "4":
                display_attendance()  # Memanggil fungsi untuk menampilkan data absensi di terminal
            elif choice == "5":
                open_attendance_with_excel()  # Memanggil fungsi untuk membuka data absensi di Excel
            elif choice == "6":
                print("Exiting...")  # Menampilkan pesan keluar
                break  # Menghentikan loop, keluar dari program
            else:
                print("Invalid choice. Try again.")  # Menampilkan pesan jika input tidak valid