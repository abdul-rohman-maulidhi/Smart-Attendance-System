import cv2
import os
import csv
import numpy as np
import pandas as pd
import sys
import pyttsx3
import speech_recognition as sr
import skfuzzy as fuzz
import skfuzzy.control as ctrl
import pyttsx3
import speech_recognition as sr

from tabulate import tabulate
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# Direktori penyimpanan
DATASET_DIR = "./data/face_dataset"
MODEL_DIR = "./data/trained_model"

ATTENDANCE_FILE = "./data/attendance/attendance.csv"
ATTENDANCE_FILE_XLSX = "./data/attendance_xlsx/attendance.xlsx"
USERS_FILE = "./data/users/users.csv"

# Inisialisasi Haar Cascades dan Recognizer
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye.xml")
smile_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_smile.xml")
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Variabel untuk deteksi
EYE_CLOSED_FRAMES = 2  # Ambang batas frame mata tertutup untuk satu kedipan
eye_blink_count = 0    # Counter kedipan mata
blink_counter = 0       # Counter untuk hitung mata tertutup berturut-turut
smile_detected = False # Flag untuk senyuman

# Inisialisasi Text-to-Speech
engine = pyttsx3.init()
engine.setProperty('rate', 150)  # Set kecepatan berbicara
engine.setProperty('volume', 0.9)  # Set volume

# Fungsi untuk mengucapkan teks
def speak_text(text):
    engine.say(text)
    engine.runAndWait()

# Fungsi untuk mendengarkan dan mengenali ucapan
def listen_for_phrase():
    recognizer = sr.Recognizer()
    with sr.Microphone() as source:
        print("Listening for 'I am here'...")
        try:
            recognizer.adjust_for_ambient_noise(source, duration=1)
            audio = recognizer.listen(source, timeout=10)
            recognized_text = recognizer.recognize_google(audio)
            print(f"Recognized: {recognized_text}")
            return recognized_text.lower()
        except sr.UnknownValueError:
            print("Could not understand the audio.")
        except sr.RequestError as e:
            print(f"Could not request results; {e}")
        except sr.WaitTimeoutError:
            print("Listening timed out.")
    return None

def get_user_info_by_id(user_id):
    users_file = USERS_FILE
    if not os.path.exists(users_file):
        return None

    with open(users_file, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["ID"] == user_id:
                return row["Name"], row["Email"]
    return None

# Fuzzy Logic untuk deteksi kedipan
def fuzzy_eye_blink(eyes_detected, eye_closed_frames):
    # Variabel fuzzy
    blink_duration = ctrl.Antecedent(np.arange(0, 10, 1), 'blink_duration')  # Durasi kedipan dalam frame
    blink_confidence = ctrl.Consequent(np.arange(0, 1.1, 0.1), 'blink_confidence')  # Kepercayaan kedipan

    # Membership functions
    blink_duration['short'] = fuzz.trapmf(blink_duration.universe, [0, 0, 2, 4])
    blink_duration['medium'] = fuzz.trimf(blink_duration.universe, [2, 4, 6])
    blink_duration['long'] = fuzz.trapmf(blink_duration.universe, [4, 6, 10, 10])

    blink_confidence['low'] = fuzz.trapmf(blink_confidence.universe, [0, 0, 0.3, 0.5])
    blink_confidence['medium'] = fuzz.trimf(blink_confidence.universe, [0.3, 0.5, 0.7])
    blink_confidence['high'] = fuzz.trapmf(blink_confidence.universe, [0.5, 0.7, 1.0, 1.0])

    # Aturan fuzzy
    rules = [
        ctrl.Rule(blink_duration['short'], blink_confidence['low']),
        ctrl.Rule(blink_duration['medium'], blink_confidence['medium']),
        ctrl.Rule(blink_duration['long'], blink_confidence['high']),
    ]

    # Sistem kontrol fuzzy
    system = ctrl.ControlSystem(rules)
    simulation = ctrl.ControlSystemSimulation(system)

    # Hitung kepercayaan berdasarkan input
    simulation.input['blink_duration'] = blink_counter
    simulation.compute()
    confidence = simulation.output['blink_confidence']

    return confidence

# Panggil fungsi fuzzy dalam pendeteksian kedipan
def check_eye_blink(eyes_detected, eye_closed_frames):
    global eye_blink_count, blink_counter
    if eyes_detected == 0:  # Mata tertutup
        blink_counter += 1
    else:  # Mata terbuka
        if blink_counter >= eye_closed_frames:
            confidence = fuzzy_eye_blink(eyes_detected, eye_closed_frames)
            if confidence > 0.5:  # Ambang kepercayaan kedipan
                eye_blink_count += 1  # Increment kedipan valid
        blink_counter = 0  # Reset counter setelah mata terbuka kembali

    return eye_blink_count

def fuzzy_smile_check(smiles, face_width):
    if len(smiles) > 0:  # Periksa jika ada senyuman yang terdeteksi
        smile_ratio = smiles[0][2] / face_width  # Ratio lebar senyuman terhadap lebar wajah
    else:
        smile_ratio = 0  # Jika tidak ada senyuman

    # Variabel fuzzy
    smile_size = ctrl.Antecedent(np.arange(0, 1.1, 0.1), 'smile_size')  # Ukuran senyuman relatif
    smile_confidence = ctrl.Consequent(np.arange(0, 1.1, 0.1), 'smile_confidence')  # Kepercayaan senyuman

    # Membership functions
    smile_size['small'] = fuzz.trapmf(smile_size.universe, [0, 0, 0.3, 0.5])
    smile_size['medium'] = fuzz.trimf(smile_size.universe, [0.3, 0.5, 0.7])
    smile_size['large'] = fuzz.trapmf(smile_size.universe, [0.5, 0.7, 1.0, 1.0])

    smile_confidence['low'] = fuzz.trapmf(smile_confidence.universe, [0, 0, 0.3, 0.5])
    smile_confidence['medium'] = fuzz.trimf(smile_confidence.universe, [0.3, 0.5, 0.7])
    smile_confidence['high'] = fuzz.trapmf(smile_confidence.universe, [0.5, 0.7, 1.0, 1.0])

    # Aturan fuzzy
    rules = [
        ctrl.Rule(smile_size['small'], smile_confidence['low']),
        ctrl.Rule(smile_size['medium'], smile_confidence['medium']),
        ctrl.Rule(smile_size['large'], smile_confidence['high']),
    ]

    # Sistem kontrol fuzzy
    system = ctrl.ControlSystem(rules)
    simulation = ctrl.ControlSystemSimulation(system)

    # Hitung kepercayaan berdasarkan input
    simulation.input['smile_size'] = smile_ratio
    simulation.compute()
    confidence = simulation.output['smile_confidence']

    return confidence > 0.5  # Ambang kepercayaan senyuman


# Fungsi utama untuk deteksi wajah, kedipan mata, dan senyuman
def recognize_faces_and_attendance():
    global eye_blink_count, smile_detected, blink_counter
    
    # Reset variabel untuk sesi baru
    eye_blink_count = 0
    smile_detected = False
    blink_counter = 0
    
    recognizer.read(os.path.join(MODEL_DIR, "face_recognizer.yml"))

    print("please blink 3x then smile for first verification!")
    speak_text("please blink 3x then smile for first verification!")
     
    # Load label map
    label_map = {}
    with open(os.path.join(MODEL_DIR, "labels.txt"), "r") as f:
        for line in f:
            label, user_id = line.strip().split(",")
            label_map[int(label)] = user_id

    cap = cv2.VideoCapture(0)
    print("Press 'q' to quit.")

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face_resized = cv2.resize(gray[y:y + h, x:x + w], (200, 200))
            try:
                label, confidence = recognizer.predict(face_resized)
                if confidence > 50.0:  # Confidence too low, treat as "Unknown"
                    detected_id = "Unknown"
                else:
                    detected_id = label_map.get(label, "Unknown")
            except Exception as e:
                detected_id = "Unknown"

            # Ambil data pengguna berdasarkan ID
            user_info = get_user_info_by_id(detected_id)
            if user_info:
                user_name, email = user_info
                display_text = f"{detected_id} | {user_name}"  # Format: ID | Name
            else:
                display_text = f"Unknown"

            # Tampilkan persegi wajah dan informasi
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 255), 2)
            cv2.putText(frame, display_text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, 
                        (0, 255, 0) if detected_id != "Unknown" else (0, 0, 255), 2)
            if detected_id != "Unknown":
                cv2.putText(frame, f"Confidence: {confidence:.2f}%", (x, y + h + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            # Additional checks for blinks and smiles if the face is recognized
            if detected_id != "Unknown":
                roi_gray = gray[y:y + h, x:x + w]

                # Eye detection
                eyes = eye_cascade.detectMultiScale(roi_gray, 1.1, 10, minSize=(30, 30))
                current_blink_count = check_eye_blink(len(eyes), EYE_CLOSED_FRAMES)
                cv2.putText(frame, f"Eye Blink Count: {current_blink_count}", (x, y - 40), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

                # Smile detection
                smiles = smile_cascade.detectMultiScale(roi_gray, 1.8, 20, minSize=(30, 30))
                smile_detected = fuzzy_smile_check(smiles, w)
                if smile_detected:
                    cv2.putText(frame, "Toothy Smile: Detected", (x, y - 70), 
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

                # Verification logic
                if current_blink_count >= 3 and smile_detected:
                    print("Please say 'I am here' to confirm your attendance.")
                    speak_text("Please say 'I am here' to confirm your attendance.")
                    response = listen_for_phrase()
                    if response and "i am here" in response:
                        cv2.putText(frame, "Verified", (x + w - 30, y + 30), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 3)
                        
                        # Log kehadiran dengan ID
                        with open(ATTENDANCE_FILE, "a", newline="") as file:
                            writer = csv.writer(file)
                            writer.writerow([detected_id, user_name, "Hadir", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                        print(f"{user_name} (ID: {detected_id}) marked as present.")
                        # Kirim email notifikasi
                        
                        email_subject = "Kehadiran Anda Tercatat"
                        email_body = (
                            f"Halo {user_name},\n\n"
                            f"Kehadiran Anda telah berhasil dicatat pada {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n\n"
                            f"ID: {detected_id}\n\n"
                            f"Terima kasih."
                        )
                        send_email_notification(email, email_subject, email_body)

                        print("Your attendance has been recorded, please check your email to see it.")
                        speak_text("Your attendance has been recorded, please check your email to see it.")
                        
                        
                        cap.release()
                        cv2.destroyAllWindows()
                        return  # Exit after attendance is marked
                    else:
                        print("Kehadiran Anda gagal dicatat. Silakan coba lagi.")
                        cap.release()
                        cv2.destroyAllWindows()
                        return

        # Tampilkan frame
        cv2.imshow("Smart Attendance System", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# Fungsi untuk mengumpulkan gambar wajah
def collect_images():
    # Input nama, ID, dan email di dalam fungsi
    person_id = input("Enter the person's ID (NIP/NIM): ").strip()
    person_name = input("Enter the person's name: ").strip()
    email = input("Enter the person's Email: ").strip()
    
    # Buat direktori berdasarkan ID pengguna
    cap = cv2.VideoCapture(0)
    count = 0
    person_dir = os.path.join(DATASET_DIR, person_id)
    os.makedirs(person_dir, exist_ok=True)

    # Simpan data pengguna ke dalam users.csv
    users_file = USERS_FILE
    if not os.path.exists(users_file):
        with open(users_file, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Email"])  # Header

    # Tambahkan data pengguna ke file
    with open(users_file, "a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([person_id, person_name, email])

    # Mulai pengumpulan gambar
    print(f"Collecting images for '{person_name}' (ID: {person_id}). Press 'q' to quit early.")
    while cap.isOpened() and count < 100:  # Default: 100 samples
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face = gray[y:y + h, x:x + w]
            face_resized = cv2.resize(face, (200, 200))
            file_path = os.path.join(person_dir, f"{count}.jpg")
            cv2.imwrite(file_path, face_resized)
            count += 1

            # Menampilkan progress pengumpulan gambar
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.putText(frame, f"Image {count}/100", (10, 30), 
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        cv2.imshow("Collecting Faces", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):  # Keluar jika 'q' ditekan
            break

    cap.release()
    cv2.destroyAllWindows()
    print(f"Collected {count} images for '{person_name}' (ID: {person_id}').")


# Fungsi untuk melatih model pengenalan wajah
def train_recognizer():
    faces, labels = [], []
    label_map = {}

    print("Training the recognizer...")
    for label, user_id in enumerate(os.listdir(DATASET_DIR)):
        label_map[label] = user_id
        user_dir = os.path.join(DATASET_DIR, user_id)
        for file_name in os.listdir(user_dir):
            img_path = os.path.join(user_dir, file_name)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            faces.append(img)
            labels.append(label)

    recognizer.train(faces, np.array(labels))
    recognizer.write(os.path.join(MODEL_DIR, "face_recognizer.yml"))

    with open(os.path.join(MODEL_DIR, "labels.txt"), "w") as f:
        for label, user_id in label_map.items():
            f.write(f"{label},{user_id}\n")
    print("Model trained and saved successfully.")

# Fungsi untuk menampilkan attendance.csv dalam bentuk tabel
def display_attendance():
    if not os.path.exists(ATTENDANCE_FILE):
        print("Attendance file does not exist.")
        return
    
    # Membaca file CSV
    df = pd.read_csv(ATTENDANCE_FILE)
    
    # Menampilkan data dalam bentuk tabel
    print("\n--- Attendance Records ---")
    print(tabulate(df, headers='keys', tablefmt='grid', showindex=False))

def open_attendance_with_excel():
    try:
        # File CSV yang akan dibuka
        csv_file = ATTENDANCE_FILE

        # Memeriksa apakah file CSV ada
        if not os.path.exists(csv_file):
            print(f"File '{csv_file}' tidak ditemukan.")
            return

        # Membaca file CSV ke dalam DataFrame
        df = pd.read_csv(csv_file)

        # File output Excel
        excel_file = ATTENDANCE_FILE_XLSX

        # Menyimpan DataFrame ke dalam file Excel
        df.to_excel(excel_file, index=False, engine='openpyxl')

        # Membuka file Excel untuk modifikasi (auto-width)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Terapkan auto-width untuk semua kolom
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Nomor kolom (1-based index)
            for cell in col:
                try:
                    # Panjang maksimum berdasarkan isi sel
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Tambahkan padding
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Format header dengan teks tebal dan tengah
        for cell in ws[1]:  # Baris pertama (header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Simpan perubahan
        wb.save(excel_file)

        print(f"File '{excel_file}' berhasil dibuat dengan kolom rapi.")

        # Membuka file Excel
        os.system(f'start EXCEL.EXE "{excel_file}"')

    except Exception as e:
        print(f"Terjadi kesalahan: {e}")
        
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def send_email_notification(to_email, subject, body):
    # Konfigurasi email
    from_email = "abdulrohmanm1304@gmail.com"
    password = "wzqz ijmp mbyw hnmb"

    try:
        # Setup email server
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_email, password)

        # Buat pesan email
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        # Kirim email
        server.send_message(msg)
        print(f"Email berhasil dikirim ke {to_email}")
        server.quit()
    except Exception as e:
        print(f"Gagal mengirim email: {e}")

# Menu utama
if __name__ == "__main__":
    if len(sys.argv) > 1:
        action = sys.argv[1]
        if action == "collect_images":
            collect_images()
        elif action == "train_recognizer":
            train_recognizer()
        elif action == "run_attendance":
            recognize_faces_and_attendance()
        elif action == "view_attendance":
            display_attendance()
        elif action == "open_excel":
            open_attendance_with_excel()
    else:
        # Original menu-driven approach
        while True:
            print("\n--- Smart Attendance System ---")
            print("1: Collect Face Data")
            print("2: Train Recognizer")
            print("3: Run Attendance System")
            print("4: View Attendance Records in Terminal")
            print("5: Open Attendance Records in Excel")
            print("6: Exit")
            choice = input("Enter your choice: ")

            if choice == "1":
                collect_images()
            elif choice == "2":
                train_recognizer()
            elif choice == "3":
                recognize_faces_and_attendance()
            elif choice == "4":
                display_attendance()
            elif choice == "5":
                open_attendance_with_excel()
            elif choice == "6":
                print("Exiting...")
                break
            else:
                print("Invalid choice. Try again.")